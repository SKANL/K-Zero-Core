"""Tools XLSX/CSV/TSV para lectura, creación y edición por copia."""
from __future__ import annotations

import csv

from k_zero_core.core.tool_safety import resolve_safe_path
from k_zero_core.core.tools.document_common import (
    _created_result,
    _export_path,
    _rows_from_json_or_markdown,
    _theme_color,
)
from k_zero_core.services.design_md import aplicar_diseno_entregable


def analizar_xlsx(path: str, max_rows: int = 20) -> str:
    """Resume hojas y primeras filas de XLSX/CSV/TSV."""
    try:
        resolved = resolve_safe_path(path)
        if resolved.suffix.lower() in {".csv", ".tsv"}:
            delimiter = "\t" if resolved.suffix.lower() == ".tsv" else ","
            with resolved.open("r", encoding="utf-8", errors="replace", newline="") as file:
                rows = list(csv.reader(file, delimiter=delimiter))[:max_rows]
            return f"Tabla: {resolved}\nFilas mostradas: {len(rows)}\n" + "\n".join(" | ".join(row) for row in rows)

        from openpyxl import load_workbook

        wb = load_workbook(str(resolved), read_only=True, data_only=False)
        lines = [f"XLSX: {resolved}", "Hojas: " + ", ".join(wb.sheetnames)]
        for sheet_name in wb.sheetnames[:5]:
            ws = wb[sheet_name]
            lines.append(f"\n[{sheet_name}]")
            for row in ws.iter_rows(max_row=max_rows, values_only=True):
                lines.append(" | ".join("" if value is None else str(value) for value in row))
        wb.close()
        return "\n".join(lines)
    except Exception as exc:
        return f"Error al analizar XLSX/CSV: {exc}"


def crear_xlsx(datos: str, nombre_sugerido: str = "datos", design_md_path: str = "") -> str:
    """Crea un XLSX estilizado desde JSON, lista de listas o tabla Markdown."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        deliverable = aplicar_diseno_entregable(datos, "xlsx", design_md_path)
        primary = _theme_color(deliverable, "colors.primary", "#1f5f8b").strip("#")
        headers, rows, sources = _rows_from_json_or_markdown(datos)
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz"
        ws.append(headers)
        for row in rows:
            ws.append(row)

        header_fill = PatternFill("solid", fgColor=primary)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        max_column = max(ws.max_column, 1)
        max_row = max(ws.max_row, 1)
        ws.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
        if max_row >= 2:
            table = Table(displayName="TablaMatriz", ref=ws.auto_filter.ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
            ws.column_dimensions[get_column_letter(column[0].column)].width = width

        if sources:
            source_ws = wb.create_sheet("Fuentes")
            source_ws.append(["URL"])
            for url in sources:
                source_ws.append([url])
            source_ws.freeze_panes = "A2"
            source_ws["A1"].fill = header_fill
            source_ws["A1"].font = Font(color="FFFFFF", bold=True)
            source_ws.column_dimensions["A"].width = min(max(len(url) for url in sources) + 2, 80)

        output = _export_path(nombre_sugerido, ".xlsx")
        wb.save(str(output))
        return _created_result(output, "xlsx", nombre_sugerido)
    except Exception as exc:
        return f"Error al crear XLSX: {exc}"


def editar_xlsx_copia(path: str, instrucciones: str) -> str:
    """Crea una copia XLSX con hoja de notas de edición."""
    try:
        resolved = resolve_safe_path(path)
        from openpyxl import load_workbook

        wb = load_workbook(str(resolved))
        ws = wb.create_sheet("Notas edición")
        ws.append(["Instrucciones"])
        ws.append([instrucciones])
        output = _export_path(resolved.stem + "_editado", ".xlsx")
        wb.save(str(output))
        return f"Copia editada: {output}"
    except Exception as exc:
        return f"Error al editar XLSX: {exc}"
